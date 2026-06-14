import openpyxl

path = r"d:\MCP\Claude Code\mau bieu giam dinh pti\input\thong_tin_giam_dinh_xe.xlsx"
wb = openpyxl.load_workbook(path)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            print(row)
