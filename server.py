# -*- coding: utf-8 -*-
"""
PTI Giám Định - Flask Backend Server
Chạy: python server.py
Truy cập: http://localhost:5000
"""

from flask import Flask, request, jsonify, send_file, send_from_directory
import os, sys, json, subprocess, shutil
import openpyxl
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
load_dotenv()

def _add_cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET,POST,OPTIONS'
    return response

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, 'input')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
ASSETS_DIR = os.path.join(BASE_DIR, 'assets')
TEMPLATE_EXCEL = os.path.join(BASE_DIR, 'docs', 'thong_tin_giam_dinh_xe.xlsx')
INPUT_EXCEL = os.path.join(INPUT_DIR, 'thong_tin_giam_dinh_xe.xlsx')

os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

app = Flask(__name__, static_folder=BASE_DIR, static_url_path='')
app.after_request(_add_cors)

@app.route('/api/<path:path>', methods=['OPTIONS'])
def handle_options(path):
    return jsonify({}), 200

ALLOWED_IMAGE = {'jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp'}
ALLOWED_EXCEL = {'xlsx', 'xls'}
ALLOWED_PDF   = {'pdf'}

GDV_DEFAULT = [
    {'ma': 'CHINH05',  'ten': 'Nguyễn Hồng Chinh', 'sdt': '0903 210 598'},
    {'ma': 'TUYENLM',  'ten': 'Lương Minh Tuyến',   'sdt': ''},
    {'ma': 'DUYNT',    'ten': 'Nguyễn Thế Duy',     'sdt': ''},
    {'ma': 'SONTT',    'ten': 'Trần Thanh Sơn',     'sdt': ''},
    {'ma': 'VIETNT05', 'ten': 'Nguyễn Tiến Việt',   'sdt': ''},
    {'ma': 'HUONGNV',  'ten': 'Nguyễn Văn Hướng',   'sdt': ''},
    {'ma': 'TUNGHX',   'ten': 'Hoàng Xuân Tùng',    'sdt': ''},
]

SCRIPT_MAP = {
    'A':   'tao_bien_ban_gd.py',
    'BFH': 'tao_3_bien_ban.py',
    'C':   'tao_bao_cao_so_bo.py',
    'D':   'tao_bao_lanh.py',
    'E':   'tao_tbtn_ycbt.py',
    'G':   'tao_cv_ngan_hang.py',
    'I':   'tao_cham_dut_khoi_phuc.py',
}

def allowed(filename, exts):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in exts


# ── Routes ──────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/assets/<path:filename>')
def serve_assets(filename):
    return send_from_directory(ASSETS_DIR, filename)

@app.route('/input/<filename>')
def serve_input(filename):
    return send_from_directory(INPUT_DIR, secure_filename(filename))


@app.route('/api/gdv-list')
def gdv_list():
    """Trả về danh sách giám định viên từ Excel template (hoặc mặc định)."""
    try:
        wb = openpyxl.load_workbook(TEMPLATE_EXCEL)
        ws = wb['GĐV']
        gdv = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                gdv.append({'ma': str(row[0]), 'ten': str(row[1]),
                             'sdt': str(row[2]) if row[2] else ''})
        return jsonify({'gdv': gdv or GDV_DEFAULT})
    except Exception:
        return jsonify({'gdv': GDV_DEFAULT})


@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    """Nhận file Excel, lưu vào input/, trả về dữ liệu đã đọc."""
    if 'file' not in request.files:
        return jsonify({'error': 'Không có file'}), 400
    f = request.files['file']
    if not f.filename or not allowed(f.filename, ALLOWED_EXCEL):
        return jsonify({'error': 'Chỉ chấp nhận file .xlsx'}), 400

    f.save(INPUT_EXCEL)

    try:
        wb = openpyxl.load_workbook(INPUT_EXCEL)
        ws_info = wb['Thông tin']
        ws_pt   = wb['Phụ tùng']

        info = {}
        for row in ws_info.iter_rows(values_only=True):
            if row[0] and str(row[0]).startswith('{') and row[2] is not None:
                key = str(row[0]).strip('{}')
                val = str(row[2]) if str(row[2]) != 'None' else ''
                info[key] = val

        phu_tung = []
        for row in ws_pt.iter_rows(min_row=2, values_only=True):
            if row[0] and 'Phương án hợp lệ' not in str(row[0]):
                phu_tung.append({
                    'ten': str(row[0]),
                    'phuong_an': str(row[1]) if row[1] else 'Thay thế có thu hồi',
                    'sl': int(row[2]) if row[2] else 1,
                })

        gdv_list = GDV_DEFAULT
        try:
            ws_gdv = wb['GĐV']
            gdv_list = []
            for row in ws_gdv.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    gdv_list.append({'ma': str(row[0]), 'ten': str(row[1]),
                                     'sdt': str(row[2]) if row[2] else ''})
        except Exception:
            pass

        return jsonify({'ok': True, 'info': info, 'phu_tung': phu_tung, 'gdv_list': gdv_list})

    except Exception as e:
        return jsonify({'error': f'Lỗi đọc Excel: {e}'}), 500


@app.route('/api/upload-images', methods=['POST'])
def upload_images():
    """Lưu ảnh và PDF vào thư mục input/."""
    saved = []
    for key in request.files:
        f = request.files[key]
        if f.filename and allowed(f.filename, ALLOWED_IMAGE | ALLOWED_PDF):
            name = secure_filename(f.filename)
            f.save(os.path.join(INPUT_DIR, name))
            saved.append(name)
    return jsonify({'ok': True, 'saved': saved})


@app.route('/api/input-images')
def list_input_images():
    """Liệt kê ảnh hiện có trong input/."""
    images = []
    if os.path.exists(INPUT_DIR):
        for f in sorted(os.listdir(INPUT_DIR)):
            if f.lower().rsplit('.', 1)[-1] in ALLOWED_IMAGE | ALLOWED_PDF:
                images.append(f)
    return jsonify({'images': images})


@app.route('/api/save-excel', methods=['POST'])
def save_excel():
    """Ghi dữ liệu form xuống file Excel template → input/."""
    data     = request.json or {}
    info     = data.get('info', {})
    phu_tung = data.get('phu_tung', [])

    # Luôn bắt đầu từ template gốc
    shutil.copy2(TEMPLATE_EXCEL, INPUT_EXCEL)

    try:
        wb = openpyxl.load_workbook(INPUT_EXCEL)
        ws_info = wb['Thông tin']
        ws_pt   = wb['Phụ tùng']

        for row in ws_info.iter_rows():
            if row[0].value and str(row[0].value).startswith('{'):
                key = str(row[0].value).strip('{}')
                if key in info:
                    row[2].value = info[key]

        # Xóa nội dung cũ rồi ghi mới
        for row in ws_pt.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        for i, pt in enumerate(phu_tung):
            ws_pt.cell(i + 2, 1, pt.get('ten', ''))
            ws_pt.cell(i + 2, 2, pt.get('phuong_an', 'Thay thế có thu hồi'))
            ws_pt.cell(i + 2, 3, int(pt.get('sl', 1)))

        wb.save(INPUT_EXCEL)
        return jsonify({'ok': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate', methods=['POST'])
def generate():
    """Chạy script sinh công văn theo danh sách được chọn."""
    data = request.json or {}
    docs = data.get('documents', [])  # ['A', 'BFH', 'G', 'I']

    results = []
    errors  = []

    for doc in docs:
        script_name = SCRIPT_MAP.get(doc)
        if not script_name:
            errors.append(f'Không có script cho mẫu {doc}')
            continue
        script_path = os.path.join(BASE_DIR, script_name)
        if not os.path.exists(script_path):
            errors.append(f'Script {script_name} không tìm thấy')
            continue

        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            cwd=BASE_DIR,
        )

        if result.returncode == 0:
            results.append(doc)
        else:
            err_msg = (result.stderr or result.stdout or '').strip()
            errors.append(f'{doc}: {err_msg[:300]}')

    # Liệt kê file output
    files = _list_output()
    return jsonify({'ok': True, 'generated': results, 'errors': errors, 'files': files})


@app.route('/api/output-files')
def list_output_files():
    return jsonify({'files': _list_output()})


@app.route('/api/reset', methods=['POST'])
def reset():
    """Xóa toàn bộ file trong input/ và output/ để bắt đầu hồ sơ mới."""
    deleted = []
    for folder in [INPUT_DIR, OUTPUT_DIR]:
        if os.path.exists(folder):
            for fname in os.listdir(folder):
                fpath = os.path.join(folder, fname)
                try:
                    os.remove(fpath)
                    deleted.append(fname)
                except Exception:
                    pass
    return jsonify({'ok': True, 'deleted': deleted})


def _list_output():
    if not os.path.exists(OUTPUT_DIR):
        return []
    return sorted(f for f in os.listdir(OUTPUT_DIR) if f.endswith('.docx'))


@app.route('/api/download/<filename>')
def download_file(filename):
    safe = secure_filename(filename)
    return send_from_directory(OUTPUT_DIR, safe, as_attachment=True)


SCAN_PROMPT = """Đây là các ảnh tài liệu xe cơ giới Việt Nam (giấy đăng ký xe, giấy phép lái xe, màn hình hệ thống PTI, giấy chứng nhận bảo hiểm, báo giá sửa chữa). Hãy trích xuất thông tin và trả về JSON với các trường sau (bỏ trống "" nếu không tìm thấy):

{
  "bien_so_xe": "",
  "hang_xe": "",
  "dong_xe": "",
  "phien_ban": "",
  "nam_sx": "",
  "so_loai": "",
  "so_khung": "",
  "so_may": "",
  "trong_tai": "",
  "so_cho_ngoi": "",
  "giay_phep_luu_hanh": "",
  "gplh_tu_ngay": "",
  "gplh_den_ngay": "",
  "giay_phep_lai_xe": "",
  "hang_gplx": "",
  "gplx_tu_ngay": "",
  "gplx_den_ngay": "",
  "chu_xe": "",
  "dien_thoai_chu_xe": "",
  "dia_chi_chu_xe": "",
  "lai_xe": "",
  "dien_thoai_lai_xe": "",
  "dia_chi_lai_xe": "",
  "so_hop_dong": "",
  "so_gcn_bh": "",
  "gcn_bh_tu_ngay": "",
  "gcn_bh_den_ngay": "",
  "gia_tri_xe": "",
  "phi_bh": "",
  "dk_bs": "",
  "so_ho_so": "",
  "ma_giam_dinh_vien": "",
  "don_vi_quan_ly": "Phòng Giám định và Cứu hộ Quảng Ninh",
  "ngay_giam_dinh": "",
  "ngay_vao_gara": "",
  "ngay_tai_nan": "",
  "khoang_gio_tai_nan": "",
  "dia_diem_tai_nan": "",
  "dien_bien_tai_nan": "",
  "nguyen_nhan_tai_nan": "",
  "tien_tt": "",
  "ten_gara": "",
  "so_tk": "",
  "ten_ngan_hang": "",
  "dia_chi_ngan_hang": "",
  "phu_tung": [
    {"ten": "Tên phụ tùng hoặc hạng mục sửa chữa", "phuong_an": "Thay thế có thu hồi|Thay thế không thu hồi|Sửa chữa", "sl": 1}
  ]
}

Lưu ý quan trọng:
- Ngày tháng định dạng DD/MM/YYYY
- Chỉ trả về JSON thuần, không thêm text hay markdown nào khác
- Biển số xe giữ nguyên định dạng gốc (vd: 14A-894.22)
- Số tiền để nguyên số không có dấu phẩy (vd: 500000000)
- Nếu cùng một thông tin xuất hiện trên nhiều ảnh, ưu tiên ảnh màn hình hệ thống PTI
- Trường "phu_tung": trích xuất từ bảng báo giá / danh sách phụ tùng nếu có; "phuong_an" chỉ nhận 3 giá trị: "Thay thế có thu hồi", "Thay thế không thu hồi", hoặc "Sửa chữa"; nếu không tìm thấy danh sách phụ tùng thì để mảng rỗng []

Hướng dẫn đọc từng loại tài liệu:
- "Phiếu xác minh phí" (XMP / mẫu xe cơ giới PTI): chứa Số GCN BH → "so_gcn_bh"; Số hợp đồng → "so_hop_dong"; Thời hạn bảo hiểm từ/đến → "gcn_bh_tu_ngay"/"gcn_bh_den_ngay"; Phí bảo hiểm (dòng tổng phí hoặc "Phí BH") → "phi_bh"; Điều kiện bổ sung → "dk_bs"
- "Giấy chứng nhận bảo hiểm" (GCN BH): số GCN BH ở đầu trang, thời hạn hiệu lực, phí BH, họ tên chủ xe, biển số
- Màn hình hệ thống PTI tab "Tổn thất - Chi trả": số hồ sơ ("so_ho_so"), mã GĐV ("ma_giam_dinh_vien"), diễn biến tai nạn ("dien_bien_tai_nan"), tên gara ("ten_gara"), tiền thanh toán ("tien_tt")
- Màn hình hệ thống PTI tab "Thông tin giám định": thông tin xe, chủ xe, lái xe, ngày giám định, ngày vào gara
- "Giấy phép lái xe" (GPLX): tìm các nhãn sau trên thẻ: "Số/No:" → dãy số ngay sau đó → "giay_phep_lai_xe"; "Họ tên/Full name:" → tên ngay sau đó → PHẢI điền vào "lai_xe" (bắt buộc, dù tên trùng chủ xe); "Nơi cư trú/Address:" → địa chỉ ngay sau đó → "dia_chi_lai_xe"; "Hạng/Class:" → "hang_gplx"; "Hiệu lực từ ngày/Date" (dạng ngày/tháng/năm trên thẻ) → "gplx_tu_ngay"; "Có giá trị đến/Expires:" → "gplx_den_ngay"; "Ngày, tháng, năm sinh/Date of birth:" không dùng; số điện thoại lái xe không có trên GPLX nên để trống
- "Giấy đăng ký xe" / "Chứng nhận đăng ký xe": "Họ tên chủ xe" → "chu_xe"; "Địa chỉ" → "dia_chi_chu_xe"; "Biển số" → "bien_so_xe"; "Nhãn hiệu" → "hang_xe"; "Loại xe" → "dong_xe"; "Số khung" → "so_khung"; "Số máy" → "so_may"; "Năm sản xuất" → "nam_sx"; "Số chỗ ngồi" → "so_cho_ngoi"
- "Giấy chứng nhận kiểm định" (đăng kiểm): "Số phiếu" → "giay_phep_luu_hanh"; "Có giá trị đến" → "gplh_den_ngay"
"""


@app.route('/api/scan-images', methods=['POST'])
def scan_images():
    """Dùng Gemini Vision API để trích xuất thông tin từ ảnh trong input/."""
    try:
        from google import genai as google_genai
        from google.genai import types as genai_types
    except ImportError:
        return jsonify({'error': 'Thư viện google-genai chưa cài. Chạy: pip install google-genai'}), 500

    api_key = os.environ.get('GEMINI_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'Chưa cấu hình GEMINI_API_KEY trong biến môi trường'}), 500

    client = google_genai.Client(api_key=api_key)

    # Thu thập ảnh và PDF từ input/ (tối đa 20 mục)
    parts = []
    files_loaded = []
    if os.path.exists(INPUT_DIR):
        for fname in sorted(os.listdir(INPUT_DIR))[:20]:
            ext = fname.lower().rsplit('.', 1)[-1]
            fpath = os.path.join(INPUT_DIR, fname)
            if ext in ALLOWED_IMAGE:
                with open(fpath, 'rb') as fp:
                    img_bytes = fp.read()
                mime = 'image/jpeg' if ext in ('jpg', 'jpeg') else f'image/{ext}'
                parts.append(genai_types.Part.from_bytes(data=img_bytes, mime_type=mime))
                files_loaded.append(f'{fname} (ảnh)')
            elif ext == 'pdf':
                pdf_text = ''
                try:
                    import pdfplumber
                    with pdfplumber.open(fpath) as pdf:
                        pdf_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
                except Exception:
                    pass
                if pdf_text.strip():
                    parts.append(genai_types.Part.from_text(text=f'[Nội dung file PDF: {fname}]\n{pdf_text}'))
                    files_loaded.append(f'{fname} (PDF text)')
                else:
                    try:
                        import fitz  # pymupdf
                        doc = fitz.open(fpath)
                        page_count = len(doc)
                        for page in doc:
                            pix = page.get_pixmap(dpi=150)
                            parts.append(genai_types.Part.from_bytes(data=pix.tobytes('png'), mime_type='image/png'))
                        doc.close()
                        files_loaded.append(f'{fname} (PDF ảnh, {page_count} trang)')
                    except Exception:
                        files_loaded.append(f'{fname} (PDF - lỗi đọc)')

    if not parts:
        return jsonify({'error': 'Chưa có ảnh hoặc PDF nào trong thư mục input. Hãy tải file lên trước.'}), 400

    parts.append(genai_types.Part.from_text(text=SCAN_PROMPT))

    try:
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=parts,
        )
        raw = response.text.strip()
    except Exception as e:
        return jsonify({'error': f'Lỗi gọi Gemini API: {e}'}), 500

    # Tách JSON từ response
    import re
    m = re.search(r'\{[\s\S]+\}', raw)
    if not m:
        return jsonify({'error': 'Không trích xuất được JSON từ phản hồi AI', 'raw': raw[:500]}), 500
    try:
        info = json.loads(m.group())
    except Exception as e:
        return jsonify({'error': f'JSON không hợp lệ: {e}', 'raw': raw[:500]}), 500

    # Nếu lai_xe trống thì mặc định = chu_xe (chủ xe thường cũng là lái xe)
    if not info.get('lai_xe') and info.get('chu_xe'):
        info['lai_xe'] = info['chu_xe']
    if not info.get('dia_chi_lai_xe') and info.get('dia_chi_chu_xe'):
        info['dia_chi_lai_xe'] = info['dia_chi_chu_xe']

    return jsonify({'ok': True, 'info': info, 'files_loaded': files_loaded})


@app.route('/api/preview/<filename>')
def preview_file(filename):
    safe = secure_filename(filename)
    path = os.path.join(OUTPUT_DIR, safe)
    if not os.path.exists(path):
        return jsonify({'error': 'File không tồn tại'}), 404
    try:
        import mammoth
        with open(path, 'rb') as f:
            result = mammoth.convert_to_html(f)
        html = result.value
        page = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body{{font-family:'Times New Roman',serif;font-size:13pt;margin:30px 40px;line-height:1.6;color:#111}}
  table{{border-collapse:collapse;width:100%}}
  td,th{{border:1px solid #999;padding:4px 8px}}
  img{{max-width:100%}}
</style></head><body>{html}</body></html>"""
        from flask import Response
        return Response(page, mimetype='text/html; charset=utf-8')
    except ImportError:
        return Response('<p style="font-family:sans-serif;padding:20px">Thư viện <b>mammoth</b> chưa được cài. Chạy: <code>pip install mammoth</code></p>', mimetype='text/html; charset=utf-8')
    except Exception as e:
        return Response(f'<p style="font-family:sans-serif;padding:20px;color:red">Lỗi xem trước: {e}</p>', mimetype='text/html; charset=utf-8')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print('=' * 55)
    print('  PTI Giám Định - Hệ thống lập hồ sơ xe cơ giới')
    print(f'  http://localhost:{port}')
    print('=' * 55)
    app.run(debug=False, host='0.0.0.0', port=port, use_reloader=False)
