import openpyxl, shutil

shutil.copy('docs/thong_tin_giam_dinh_xe.xlsx', 'docs/thong_tin_giam_dinh_xe_filled.xlsx')

wb = openpyxl.load_workbook('docs/thong_tin_giam_dinh_xe_filled.xlsx')
ws = wb['Thông tin']

data = {
    # A. Xe — từ đăng ký xe (1.1.jpg) + kiểm định (3.1.jpg)
    '{bien_so_xe}':             '14H-042.80',
    '{hang_xe}':                'KIA',
    '{dong_xe}':                'CARNIVAL',
    '{phien_ban}':              'CARNIVAL LUXURY 2.2D 8 CHỖ',
    '{nam_sx}':                 '2023',
    '{so_loai}':                'KA4 2.2 AT FL8',
    '{so_khung}':               'RNYND5BA8PC321432',
    '{so_may}':                 'D4HEPH601440',
    '{trong_tai}':              '',
    '{so_cho_ngoi}':            '8',

    # B. Giấy tờ — từ kiểm định (3.1.jpg) + GPLX (2.2.jpg)
    '{giay_phep_luu_hanh}':     'VA 3365650',
    '{gplh_tu_ngay}':           '',
    '{gplh_den_ngay}':          '',
    '{giay_phep_lai_xe}':       'CA400350',
    '{hang_gplx}':              'B1',
    '{gplx_tu_ngay}':           '25/02/2011',
    '{gplx_den_ngay}':          '',

    # C. Chủ xe / Lái xe — từ đăng ký xe (1.1.jpg) + CCCD lái xe (1.4.jpg) + GCN BH
    '{chu_xe}':                 'CÔNG TY TNHH PHÚC XUYÊN',
    '{dien_thoai_chu_xe}':      '0203.663366',
    '{dia_chi_chu_xe}':         'Tổ 7, Khu 1, Phường Yên Thanh, Thành phố Uông Bí, Tỉnh Quảng Ninh',
    '{lai_xe}':                 'ĐOÀN THỂ XUYÊN',
    '{dien_thoai_lai_xe}':      '',
    '{dia_chi_lai_xe}':         'Tổ 1, Khu 1, Phường Yên Thanh, Thành phố Uông Bí, Tỉnh Quảng Ninh',

    # D. Hợp đồng BH — từ GCN BH + Phiếu XMP
    '{so_hop_dong}':            '0000277/HD/013-013/XO/2025',
    '{so_gcn_bh}':              '013OTTN+250004310',
    '{gcn_bh_tu_ngay}':         '28/12/2025',
    '{gcn_bh_den_ngay}':        '28/12/2026',
    '{gia_tri_xe}':             '970,000,000',
    '{phi_bh}':                 '14,841,000',
    '{dk_bs}':                  'BS02, BS05, BS06',

    # E. Hồ sơ
    '{so_ho_so}':               '...',
    '{ma_giam_dinh_vien}':      '...',
    '{don_vi_quan_ly}':         'Phòng Giám định và Cứu hộ Quảng Ninh',
    '{ngay_giam_dinh}':         '...',
    '{ngay_vao_gara}':          '...',

    # F. Tai nạn
    '{ngay_tai_nan}':           '08/06/2026',
    '{khoang_gio_tai_nan}':     '...',
    '{dia_diem_tai_nan}':       '...',
    '{dien_bien_tai_nan}':      '...',
    '{nguyen_nhan_tai_nan}':    '...',

    # G. Tài chính
    '{ngay_hoa_don}':           '...',
    '{thang_hoa_don}':          '...',
    '{nam_hoa_don}':            '...',
    '{tong_thanh_toan_hoa_don}':'...',

    # H. Gara
    '{ten_gara}':               '...',
    '{so_tk}':                  '...',
    '{ten_ngan_hang}':          '...',
    '{dia_chi_ngan_hang}':      '...',
}

filled = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.column == 1 and cell.value in data:
            ws.cell(row=cell.row, column=3).value = data[cell.value]
            filled += 1

wb.save('docs/thong_tin_giam_dinh_xe_filled.xlsx')
print(f'Đã điền {filled} trường vào cột C.')
print()
print('=== CÒN THIẾU (cần bổ sung) ===')
missing = [k for k, v in data.items() if v == '']
for m in missing:
    print(f'  {m}')
